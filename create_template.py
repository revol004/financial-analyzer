import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import numpy as np

# Categories provided
categories = [
    "total_assets",
    "longterm_assets",
    "current_assets",
    "current_receivables",
    "cash_&_equivalents",
    "marketable_securities",
    "equity",
    "non_controlling_interest",
    "longterm_liabilities",
    "longterm_loans_&_borrowings_liabilities",
    "longterm_bonds_liabilities",
    "longterm_leases_liabilities",
    "longterm_other_financial_liabilities",
    "current_liabilities",
    "current_loans_&_borrowings_liabilities",
    "current_bonds__liabilities",
    "current_leases__liabilities",
    "current_other_financial__liabilities",
    "accruals",
    "sales",
    "ebit",
    "net_income",
    "operating_cash_flows",
    "capex",
    "financial_cash_flows",
    "depreciation",
    "outstanding_shares",
    "price",
]

# Create data for 2023 and 2022, 4 quarters each
data = []
for year in [2023, 2022]:
    for quarter in [1, 2, 3, 4]:
        row = {"year": year, "quarter": quarter}
        # Generate sample data - realistic proportions
        base_multiplier = 1.05 if year == 2023 else 1.0

        # Assets (in millions)
        row["total_assets"] = 5000000 * base_multiplier
        row["longterm_assets"] = 3000000 * base_multiplier
        row["current_assets"] = 2000000 * base_multiplier
        row["current_receivables"] = 800000 * base_multiplier
        row["cash_&_equivalents"] = 600000 * base_multiplier
        row["marketable_securities"] = 200000 * base_multiplier

        # Liabilities & Equity
        row["equity"] = 3000000 * base_multiplier
        row["non_controlling_interest"] = 100000 * base_multiplier
        row["longterm_liabilities"] = 1200000 * base_multiplier
        row["longterm_loans_&_borrowings_liabilities"] = 600000 * base_multiplier
        row["longterm_bonds_liabilities"] = 400000 * base_multiplier
        row["longterm_leases_liabilities"] = 100000 * base_multiplier
        row["longterm_other_financial_liabilities"] = 100000 * base_multiplier
        row["current_liabilities"] = 700000 * base_multiplier
        row["current_loans_&_borrowings_liabilities"] = 300000 * base_multiplier
        row["current_bonds__liabilities"] = 200000 * base_multiplier
        row["current_leases__liabilities"] = 100000 * base_multiplier
        row["current_other_financial__liabilities"] = 100000 * base_multiplier
        row["accruals"] = 250000 * base_multiplier

        # P&L (quarterly)
        row["sales"] = 250000 * base_multiplier * (1 + np.random.uniform(-0.1, 0.1))
        row["ebit"] = row["sales"] * 0.25
        row["net_income"] = row["sales"] * 0.15

        # Cash Flows
        row["operating_cash_flows"] = row["net_income"] * 1.2
        row["capex"] = 150000 * base_multiplier
        row["financial_cash_flows"] = 50000 * base_multiplier
        row["depreciation"] = 80000 * base_multiplier

        # Other
        row["outstanding_shares"] = 100000000
        row["price"] = 50 + np.random.uniform(-5, 5)

        data.append(row)

# Create DataFrame
df = pd.DataFrame(data)

# Save to XLSX
output_path = (
    r"c:\Users\jakub\Documents\financial-analyzer\template_import_quarters.xlsx"
)
df.to_excel(output_path, index=False, sheet_name="Financial Data")

# Format the XLSX
wb = load_workbook(output_path)
ws = wb.active

# Header formatting
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 20)
    ws.column_dimensions[column_letter].width = adjusted_width

# Format numeric columns
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        col_letter = cell.column_letter
        col_name = ws[f"{col_letter}1"].value

        if col_name in ["year", "quarter", "outstanding_shares"]:
            cell.number_format = "0"
        elif col_name == "price":
            cell.number_format = "0.00"
        else:
            cell.number_format = "#,##0"

wb.save(output_path)
print(f"✅ XLSX template created: {output_path}")
print(
    f"📊 Columns: {len(categories) + 2} (year, quarter, + {len(categories)} financial variables)"
)
print(f"📈 Rows: {len(data)} (2 years × 4 quarters)")
